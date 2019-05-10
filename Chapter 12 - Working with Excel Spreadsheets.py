import os

import openpyxl, glob
from openpyxl.styles import Font

# Multiplication Table Maker
# Create a program multiplicationTable.py that takes a number N 
# from the command line and creates an N×N multiplication table 
# in an Excel spreadsheet.


def multi_Table_Maker(num):
        
    wb = openpyxl.Workbook()
    ws = wb.active

    #populate col A and row 1 with bold font
    boldFont = Font(bold=True)
    for i in range(1,num+1):
        ws.cell(1,i+1).value = i
        ws.cell(1,i+1).font = boldFont
        ws.cell(i+1,1).value = i
        ws.cell(i+1,1).font= boldFont

    #insert multipication formula in between
    for i in range(1,num+1): #row
        for j in range(1,num+1): #col
            ws.cell(i+1,j+1).value = ws.cell(i+1,1).value * ws.cell(1,j+1).value

    wb.save('test.xlsx')

# Blank Row Inserter
# Create a program blankRowInserter.py that takes two integers and 
# a filename string as command line arguments. Let’s call the first 
# integer N and the second integer M. Starting at row N, the program 
# should insert M blank rows into the spreadsheet.

def blank_Row_Inserter(n,m, path):
    os.chdir(path)
    wb = openpyxl.load_workbook('test.xlsx')
    ws1 = wb.active
    lastRow = ws1.max_row +1
    lastCol = ws1.max_column +1

    ws2 = wb.create_sheet('Sheet2')
    for i in range(1, lastRow):
        for j in range(1, lastCol):
            if i>=n: 
                ws2.cell(i+m,j).value = ws1.cell(i,j).value 
            else:
                ws2.cell(i,j).value = ws1.cell(i,j).value
    wb.save('test_new.xlsx')

# Spreadsheet Cell Inverter
# Write a program to invert the row and column of the cells in 
# the spreadsheet. For example, the value at row 5, column 3 will 
# be at row 3, column 5 (and vice versa). This should be done for 
# all cells in the spreadsheet. For example, the “before” and “after” 
# spreadsheets would look something like Figure 12-13.
def cell_Inverter(path):
    os.chdir(path)
    wb = openpyxl.load_workbook('test.xlsx')
    ws1 = wb.active
    lastRow = ws1.max_row +1
    lastCol = ws1.max_column +1

    ws2 = wb.create_sheet('Sheet2')
    for i in range(1, lastRow):
        for j in range(1, lastCol):
            ws2.cell(j,i).value = ws1.cell(i,j).value

    wb.save('test_new.xlsx')

# Text Files to Spreadsheet
# Write a program to read in the contents of several text files (you can 
# make the text files yourself) and insert those contents into a spreadsheet,
#  with one line of text per row. The lines of the first text file will be in the 
# cells of column A, the lines of the second text file will be in the cells of 
# column B, and so on.
def txtToWs(path):
    os.chdir(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    # lastRow = ws.max_row +1
    # lastCol = ws.max_column +1
    fileList = glob.glob(os.path.join(path,'*.txt'))
    j = 1 # iternate one new column for each txt file
    for file in fileList:
        with open(file,'r') as f:
            lines = f.readlines()
            i=1
            for line in lines:
                print(line)
                ws.cell(i,j).value = line
                i+=1
        j+=1
    wb.save('test_new.xlsx')
# Spreadsheet to Text Files
# Write a program that performs the tasks of the previous program in reverse order: 
# The program should open a spreadsheet and write the cells of column A 
# into one text file, the cells of column B into another text file, and so on.
def wsToTxt(path):
    os.chdir(path)
    wb = openpyxl.load_workbook('test_new.xlsx')
    ws = wb.active
    lastRow = ws.max_row +1
    lastCol = ws.max_column +1
 
    for i in range(1, lastCol):
        with open('xlsExport_'+str(i)+'.txt', 'a') as f:
            for j in range(1, lastRow):
                f.write(str(ws.cell(j,i).value))
                j+=1
        i+=1

if __name__ == "__main__":
    path = r'C:\Users\Fung\Documents\vsc git repository\Automating-Boring-Stuff-with-Python-projects'
    while True:
        try:
            num = int(input('Please input a positive integer to make the multiplcation table:'))
            if num > 0:  #num must be larger than 0
                multi_Table_Maker(num)
                break
        except Exception:
            pass
    blank_Row_Inserter(3, 2, path)
    cell_Inverter(path)
    txtToWs(r'C:\Users\Fung\Downloads\testDoc')
    wsToTxt(r'C:\Users\Fung\Downloads\testDoc')

    