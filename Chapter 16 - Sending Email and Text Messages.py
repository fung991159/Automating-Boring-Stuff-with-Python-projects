# import smtplib

#Random Chore Assignment Emailer

import os, datetime, random
import smtplib, imapclient, pyzmail

import openpyxl
def assignment_Emailer():
    #open excel file, get e-mail address and last chore
    wb = openpyxl.load_workbook('Chapter16_choresTable.xlsx')
    ws = wb.active
    lastCol = ws.max_column
    lastRow = ws.max_row
    #If today is larger than the last assiged date, then random assign chore without repeating last chore
    currentDate =datetime.datetime.today().date()
    lastDate = ws.cell(1,lastCol).value.date()

    if lastDate >= currentDate: 
        print('Chores has been assigned and mail sent, there is no need to do it again today!')
    else:
        print('It is time to assign new chores for everybody!')    
        chores=[]
        for i in range(2,lastRow+1):
            if ws.cell(i,lastCol).value not in chores:
                chores.append(ws.cell(i,lastCol).value) # assume previous assignment is unique
        
        #assign chores without repetition from last time
        print('assigning chores')
        i = 2
        while i <= lastRow:
            randomChore = random.choice(chores)
            if randomChore == ws.cell(i,lastCol).value:
                i = i #restart the loop
            else:
                chores.remove(randomChore)
                ws.cell(i,lastCol+1).value = randomChore
                i +=1
        ws.cell(1,lastCol+1).value = currentDate  #assign current date to header
        ws.cell(1,lastCol+1).number_format = ws.cell(1,lastCol).number_format #copy last date number format

        #list of lists to hold information
        emailInfo = []  #creating a list is probably not necessary, but just for practice :)
        for i in range(2,ws.max_row+1):
            emailInfo.append([ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,lastCol).value])
        
        #Gather email element
        print('reading e-mails info')
        for i in range(len(emailInfo)):
            receipant = emailInfo[i][0]
            eMailAdd = emailInfo[i][1]
            assigedChore = emailInfo[i][2]

            #E-mail each person their assigned chores
            smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login('@gmail.com', 'pwd')    # <<<< Login email address and password
            eMailSubject= f'Subject: {receipant}, your chore today was assigned! !\n This time you get to {assigedChore}'
            print(f'Sending mail to {receipant} {eMailAdd}')
            smtpObj.sendmail('@gmail.com', eMailAdd,eMailSubject) #<<<<Login email address
        smtpObj.quit()
        wb.save('Chapter16_choresTable.xlsx') #only save excel file after everything is done
        print('All notification mail sent!')

# Auto Unsubscriber
def auto_Unsubsubscriber():
    targetDate = datetime.datetime.now() - datetime.timedelta(days=10)
    targetDateStr = targetDate.strftime('%d-%b-%Y')
    imapObj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
    imapObj.login('test@gmail.com', 'pwd')    # gmail login and pwd here
    imapObj.select_folder('INBOX', readonly=True)
    UIDs = imapObj.search(['SINCE', targetDateStr])
    unsubscribeLinkUrls = [] #collect all links here
    for UID in UIDs:
        rawMessages = imapObj.fetch(UID, ['BODY[]', 'FLAGS'])
        import pyzmail, bs4
        message = pyzmail.PyzMessage.factory(rawMessages[UID][b'BODY[]'])
        if message.html_part != None:  #read mail that have html format available
            soupSearch = bs4.BeautifulSoup(message.html_part.get_payload().decode(message.html_part.charset),'html.parser')
            linkList = soupSearch.select('a')
            for link in linkList:
                if 'unsubscribe' in link:
                    unsubscribeLinkUrls.append(link.get('href'))  #get unsubscribe link
    
    #open each unsubscribe link in webbrowser
    import webbrowser
    for unsubscribeUrl in unsubscribeLinkUrls:
        webbrowser.open(unsubscribeUrl)

def control_Computer_by_mail():
    #log any error, if any
    import logging 
    logging.basicConfig(filename = 'control_Computer_by_mail.txt', level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')
    try:
        #check mail from my email address with pwd
        print('reading E-mails for torrent links')
        today = datetime.datetime.now().strftime('%d-%b-%Y')
        imapObj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
        myEmailAddr = 'abc@gmail.com'        #login acct here
        myEmailPwd = 'pwd'                   #login pwd here
        imapObj.login(myEmailAddr, myEmailPwd)    # gmail login and pwd here
        imapObj.select_folder('INBOX', readonly=False)
        UIDs = imapObj.search(['ON', today, 'FROM', myEmailAddr])   #get only message from myself
        emailVerificationPwd = 'ye8sd'
        #gather content BitTorrent Link and delete mail
        torrentLinks = []
        for UID in UIDs:
            rawMessages = imapObj.fetch(UID, ['BODY[]', 'FLAGS'])
            import pyzmail, bs4
            message = pyzmail.PyzMessage.factory(rawMessages[UID][b'BODY[]'])
            if message.get_subject() != emailVerificationPwd:
                pass
            else:
                if message.html_part != None:  #read mail that have html format available
                    soupSearch = bs4.BeautifulSoup(message.html_part.get_payload().decode(message.html_part.charset),'html.parser')
                    torrentLinkElement = soupSearch.select('div')
                    torrentLinks.append(torrentLinkElement[0].text)
                    imapObj.delete_messages(UIDs) #delete mail
        
        if torrentLinks == []:
            print('no mail found')
            pass #do nothing if no E_mail found
        else:
            #open link and download file in qtorrent
            import subprocess
            print('E-mail(s), opening torrent links in qbTorret')
            for torrentLink in torrentLinks:
                qbProcess = subprocess.Popen(['C:\Program Files\qBittorrent\qbittorrent.exe', torrentLink])
            #when download finish and qbitorrent quit, send email confirm 
            qbProcess.wait()
            #E-mail each person their assigned chores
            print('download completed: sending e-Mail to self')
            smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login(myEmailAddr, myEmailPwd)    
            eMailSubject= f'Subject: qbTorrent download completed! \n Download completed!'
            smtpObj.sendmail(myEmailAddr, myEmailAddr,eMailSubject) 
            smtpObj.quit()
    except Exception as err:
        print('critial error found! logging to text file')
        with open('control_Computer_by_mail.txt', 'a') as f:
            f.write(str(err))
    


if __name__ == "__main__":
    # assignment_Emailer()
    # auto_Unsubsubscriber()
    control_Computer_by_mail()
